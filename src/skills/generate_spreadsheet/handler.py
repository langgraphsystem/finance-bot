"""Generate Excel spreadsheets via LLM-authored openpyxl code in E2B sandbox."""

import asyncio
import logging
import re
from typing import Any

from src.core.context import SessionContext
from src.core.llm.clients import generate_text
from src.core.observability import observe
from src.gateway.types import IncomingMessage
from src.skills._i18n import register_strings, t_cached
from src.skills.base import SkillResult

logger = logging.getLogger(__name__)

_STRINGS = {
    "en": {
        "no_description": (
            "What kind of spreadsheet should I create? Describe the data and structure."
        ),
        "spreadsheet_ready": "<b>{filename}</b> — your spreadsheet is ready.",
        "generation_failed": "Failed to generate spreadsheet. Try a simpler description.",
    },
    "ru": {
        "no_description": ("Какую таблицу создать? Опишите данные и структуру."),
        "spreadsheet_ready": "<b>{filename}</b> — ваша таблица готова.",
        "generation_failed": "Не удалось создать таблицу. Попробуйте более простое описание.",
    },
    "es": {
        "no_description": (
            "Que tipo de hoja de calculo debo crear? Describa los datos y la estructura."
        ),
        "spreadsheet_ready": "<b>{filename}</b> — su hoja de calculo esta lista.",
        "generation_failed": (
            "No se pudo generar la hoja de calculo. "
            "Pruebe una descripcion mas simple."
        ),
    },
}
register_strings("generate_spreadsheet", _STRINGS)

SPREADSHEET_SYSTEM_PROMPT = """\
You are an Excel spreadsheet generator. You write Python code using openpyxl
that creates an .xlsx file at /tmp/output.xlsx.

Rules:
- Always start with: from openpyxl import Workbook
- Always save to /tmp/output.xlsx
- Use openpyxl.styles for formatting (Font, PatternFill, Alignment, Border, Side)
- Add headers with bold font and colored fill
- Auto-adjust column widths based on content
- Add sample data rows when the user doesn't provide specific data
- Use number formats for currency/percentage cells
- Respond ONLY with Python code, no explanations outside the code
- The code must run standalone without errors
- Do NOT use any library except openpyxl and its submodules"""

FALLBACK_SYSTEM_PROMPT = """\
You are an Excel spreadsheet generator. Describe the spreadsheet structure as JSON.
Return a JSON object with:
{
  "title": "Sheet title",
  "headers": ["Column A", "Column B", ...],
  "rows": [["val1", "val2", ...], ...],
  "column_widths": [15, 20, ...]
}
Respond ONLY with valid JSON."""


def _strip_markdown_fences(text: str) -> str:
    """Remove ```language ... ``` wrappers from LLM output."""
    text = text.strip()
    match = re.match(r"^```\w*\n(.*?)```$", text, re.DOTALL)
    if match:
        return match.group(1).strip()
    return text


async def _build_fallback_xlsx(description: str) -> bytes | None:
    """Generate a spreadsheet without E2B using LLM JSON + local openpyxl."""
    import io
    import json

    spec_text = await generate_text(
        model="claude-sonnet-4-6",
        system=FALLBACK_SYSTEM_PROMPT,
        prompt=f"Create a spreadsheet: {description}",
        max_tokens=2048,
    )
    spec_text = _strip_markdown_fences(spec_text)

    try:
        spec = json.loads(spec_text)
    except json.JSONDecodeError:
        logger.warning("Failed to parse spreadsheet JSON spec")
        return None

    def _create():
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = spec.get("title", "Sheet1")

        headers = spec.get("headers", [])
        rows = spec.get("rows", [])
        col_widths = spec.get("column_widths", [])

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_align = Alignment(horizontal="center")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        if not col_widths:
            for col_idx in range(1, len(headers) + 1):
                max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 8
                for row_idx in range(2, len(rows) + 2):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val:
                        max_len = max(max_len, len(str(cell_val)))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                    max_len + 4, 50
                )

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    return await asyncio.to_thread(_create)


class GenerateSpreadsheetSkill:
    name = "generate_spreadsheet"
    intents = ["generate_spreadsheet"]
    model = "claude-sonnet-4-6"

    @observe(name="generate_spreadsheet")
    async def execute(
        self,
        message: IncomingMessage,
        context: SessionContext,
        intent_data: dict[str, Any],
    ) -> SkillResult:
        description = (intent_data.get("description") or message.text or "").strip()
        lang = context.language or "en"

        if not description:
            return SkillResult(
                response_text=t_cached(_STRINGS, "no_description", lang, "generate_spreadsheet")
            )

        # Try E2B sandbox first
        try:
            from src.tools.e2b_file_utils import execute_code_with_file

            code = await generate_text(
                model=self.model,
                system=SPREADSHEET_SYSTEM_PROMPT,
                prompt=f"Create a spreadsheet: {description}",
                max_tokens=4096,
            )
            code = _strip_markdown_fences(code)

            file_bytes, stdout = await execute_code_with_file(
                code=code,
                output_filename="output.xlsx",
                language="python",
                timeout=60,
                install_deps=["openpyxl"],
            )

            if file_bytes:
                filename = _make_filename(description)
                logger.info(
                    "Spreadsheet generated via E2B for user %s (%d bytes)",
                    context.user_id,
                    len(file_bytes),
                )
                await _store_episode(context, description, len(file_bytes), "e2b")
                return SkillResult(
                    response_text=t_cached(
                        _STRINGS, "spreadsheet_ready", lang, "generate_spreadsheet"
                    ).format(filename=filename),
                    document=file_bytes,
                    document_name=filename,
                )

            logger.warning("E2B produced no file, falling back to local: %s", stdout)
        except Exception as e:
            logger.warning("E2B spreadsheet generation failed, using fallback: %s", e)

        # Fallback: local openpyxl via LLM JSON spec
        file_bytes = await _build_fallback_xlsx(description)
        if file_bytes:
            filename = _make_filename(description)
            logger.info(
                "Spreadsheet generated locally for user %s (%d bytes)",
                context.user_id,
                len(file_bytes),
            )
            await _store_episode(context, description, len(file_bytes), "local")
            return SkillResult(
                response_text=t_cached(
                    _STRINGS, "spreadsheet_ready", lang, "generate_spreadsheet"
                ).format(filename=filename),
                document=file_bytes,
                document_name=filename,
            )

        return SkillResult(
            response_text=t_cached(
                _STRINGS, "generation_failed", lang, "generate_spreadsheet"
            )
        )

    def get_system_prompt(self, context: SessionContext) -> str:
        return SPREADSHEET_SYSTEM_PROMPT


async def _store_episode(
    context: SessionContext, description: str, size: int, method: str
) -> None:
    """Store episodic memory after successful generation."""
    try:
        from src.core.memory.episodic import store_episode

        await store_episode(
            user_id=str(context.user_id),
            family_id=str(context.family_id),
            intent="generate_spreadsheet",
            result_metadata={
                "format": "xlsx",
                "description": description[:100],
                "size": size,
                "method": method,
            },
        )
    except Exception as e:
        logger.debug("Episode storage failed: %s", e)


def _make_filename(description: str) -> str:
    """Generate a short filename from the description."""
    slug = re.sub(r"[^a-zA-Z0-9\s]", "", description.lower())
    slug = "_".join(slug.split()[:5])
    if not slug:
        slug = "spreadsheet"
    if len(slug) > 40:
        slug = slug[:40].rstrip("_")
    return f"{slug}.xlsx"


skill = GenerateSpreadsheetSkill()
