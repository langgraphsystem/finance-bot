"""Export Excel — generate .xlsx spreadsheets from user data."""

import io
import logging
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.core.context import SessionContext
from src.core.observability import observe
from src.gateway.types import IncomingMessage
from src.skills._i18n import register_strings, t_cached
from src.skills.base import SkillResult
from src.skills.prompt_loader import load_prompt

logger = logging.getLogger(__name__)

_STRINGS = {
    "en": {
        "no_expenses": "No expenses found for this period.",
        "exported_expenses": "Exported {count} expenses ({currency} {total}).",
        "no_tasks": "No tasks found.",
        "exported_tasks": "Exported {count} tasks.",
        "no_contacts": "No contacts found.",
        "exported_contacts": "Exported {count} contacts.",
        "transactions_sheet": "Transactions",
        "date": "Date",
        "merchant": "Merchant",
        "category": "Category",
        "amount": "Amount",
        "description": "Description",
        "summary_sheet": "Summary",
        "metric": "Metric",
        "value": "Value",
        "period": "Period",
        "total": "Total",
        "transactions": "Transactions",
        "tasks_sheet": "Tasks",
        "title": "Title",
        "status": "Status",
        "due_date": "Due Date",
        "created": "Created",
        "contacts_sheet": "Contacts",
        "name": "Name",
        "phone": "Phone",
        "email": "Email",
        "role": "Role",
        "notes": "Notes",
    },
    "ru": {
        "no_expenses": "По этому периоду трат не найдено.",
        "exported_expenses": "Экспортировано {count} расход(ов) ({currency} {total}).",
        "no_tasks": "Задачи не найдены.",
        "exported_tasks": "Экспортировано {count} задач(и).",
        "no_contacts": "Контакты не найдены.",
        "exported_contacts": "Экспортировано {count} контакт(ов).",
        "transactions_sheet": "Операции",
        "date": "Дата",
        "merchant": "Магазин",
        "category": "Категория",
        "amount": "Сумма",
        "description": "Описание",
        "summary_sheet": "Итоги",
        "metric": "Метрика",
        "value": "Значение",
        "period": "Период",
        "total": "Итого",
        "transactions": "Операций",
        "tasks_sheet": "Задачи",
        "title": "Название",
        "status": "Статус",
        "due_date": "Срок",
        "created": "Создано",
        "contacts_sheet": "Контакты",
        "name": "Имя",
        "phone": "Телефон",
        "email": "Электронная почта",
        "role": "Должность",
        "notes": "Заметки",
    },
    "es": {
        "no_expenses": "No se encontraron gastos para este período.",
        "exported_expenses": "Se exportaron {count} gastos ({currency} {total}).",
        "no_tasks": "No se encontraron tareas.",
        "exported_tasks": "Se exportaron {count} tareas.",
        "no_contacts": "No se encontraron contactos.",
        "exported_contacts": "Se exportaron {count} contactos.",
        "transactions_sheet": "Transacciones",
        "date": "Fecha",
        "merchant": "Comerciante",
        "category": "Categoría",
        "amount": "Cantidad",
        "description": "Descripción",
        "summary_sheet": "Resumen",
        "metric": "Métrica",
        "value": "Valor",
        "period": "Período",
        "total": "Total",
        "transactions": "Transacciones",
        "tasks_sheet": "Tareas",
        "title": "Título",
        "status": "Estado",
        "due_date": "Fecha de vencimiento",
        "created": "Creado",
        "contacts_sheet": "Contactos",
        "name": "Nombre",
        "phone": "Teléfono",
        "email": "Correo electrónico",
        "role": "Función",
        "notes": "Notas",
    },
}
register_strings("export_excel", _STRINGS)

_DEFAULT_SYSTEM_PROMPT = """\
You help the user export their data as an Excel spreadsheet.
Supported export types: expenses, tasks, contacts.
Respond in: {language}."""

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")


def _auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)


def _style_headers(ws):
    """Apply bold font and blue fill to the first row."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _detect_export_type(intent_data: dict, message_text: str) -> str:
    """Detect export type from intent_data or message keywords."""
    if intent_data.get("export_type"):
        return intent_data["export_type"]
    text_lower = (message_text or "").lower()
    if any(kw in text_lower for kw in ("task", "задач", "todo", "дела")):
        return "tasks"
    if any(kw in text_lower for kw in ("contact", "контакт", "client", "клиент")):
        return "contacts"
    return "expenses"


def _parse_date_range(intent_data: dict) -> tuple[date, date]:
    """Parse date range from intent_data, default to current month."""
    period = intent_data.get("period", "")
    today = date.today()

    if period == "week":
        return today - timedelta(days=7), today
    if period == "year":
        return date(today.year, 1, 1), today

    if intent_data.get("date_from") and intent_data.get("date_to"):
        try:
            return (
                date.fromisoformat(intent_data["date_from"]),
                date.fromisoformat(intent_data["date_to"]),
            )
        except ValueError:
            pass

    # Default: current month
    return date(today.year, today.month, 1), today


class ExportExcelSkill:
    name = "export_excel"
    intents = ["export_excel"]
    model = "claude-haiku-4-5"

    @observe(name="export_excel")
    async def execute(
        self,
        message: IncomingMessage,
        context: SessionContext,
        intent_data: dict[str, Any],
    ) -> SkillResult:
        export_type = _detect_export_type(intent_data, message.text or "")

        if export_type == "tasks":
            return await self._export_tasks(context)
        if export_type == "contacts":
            return await self._export_contacts(context)
        return await self._export_expenses(context, intent_data)

    async def _export_expenses(self, context: SessionContext, intent_data: dict) -> SkillResult:
        from sqlalchemy import text

        from src.core.db import async_session

        date_from, date_to = _parse_date_range(intent_data)

        async with async_session() as session:
            result = await session.execute(
                text("""
                    SELECT date, merchant, category, amount, description
                    FROM transactions
                    WHERE family_id = :fid AND type = 'expense'
                      AND date >= :start AND date <= :end
                    ORDER BY date DESC
                """),
                {"fid": context.family_id, "start": date_from, "end": date_to},
            )
            rows = result.all()

        if not rows:
            return SkillResult(
                response_text=t_cached(
                    _STRINGS,
                    "no_expenses",
                    context.language or "en",
                    namespace="export_excel",
                )
            )

        wb = Workbook()
        ws = wb.active
        ws.title = t_cached(
            _STRINGS,
            "transactions_sheet",
            context.language or "en",
            namespace="export_excel",
        )
        ws.append(
            [
                t_cached(_STRINGS, "date", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "merchant", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "category", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "amount", context.language or "en", namespace="export_excel"),
                t_cached(
                    _STRINGS, "description", context.language or "en", namespace="export_excel"
                ),
            ]
        )
        _style_headers(ws)

        total = 0
        cat_totals: dict[str, float] = {}
        for r in rows:
            ws.append(
                [
                    str(r.date) if r.date else "",
                    r.merchant or "",
                    r.category or "",
                    float(r.amount),
                    r.description or "",
                ]
            )
            total += float(r.amount)
            cat = r.category or "Other"
            cat_totals[cat] = cat_totals.get(cat, 0) + float(r.amount)
        _auto_fit_columns(ws)

        # Summary sheet
        ws2 = wb.create_sheet(
            t_cached(_STRINGS, "summary_sheet", context.language or "en", namespace="export_excel")
        )
        ws2.append(
            [
                t_cached(_STRINGS, "metric", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "value", context.language or "en", namespace="export_excel"),
            ]
        )
        _style_headers(ws2)
        ws2.append(
            [
                t_cached(_STRINGS, "period", context.language or "en", namespace="export_excel"),
                f"{date_from} — {date_to}",
            ]
        )
        ws2.append(
            [
                t_cached(_STRINGS, "total", context.language or "en", namespace="export_excel"),
                f"{context.currency} {total:.2f}",
            ]
        )
        ws2.append(
            [
                t_cached(
                    _STRINGS, "transactions", context.language or "en", namespace="export_excel"
                ),
                len(rows),
            ]
        )
        ws2.append([])
        ws2.append(
            [
                t_cached(_STRINGS, "category", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "amount", context.language or "en", namespace="export_excel"),
            ]
        )
        for cat_name in sorted(cat_totals, key=cat_totals.get, reverse=True):
            ws2.append([cat_name, f"{context.currency} {cat_totals[cat_name]:.2f}"])
        _auto_fit_columns(ws2)

        buf = io.BytesIO()
        wb.save(buf)

        fname = f"expenses_{date_from}_{date_to}.xlsx"
        return SkillResult(
            response_text=t_cached(
                _STRINGS,
                "exported_expenses",
                context.language or "en",
                namespace="export_excel",
                count=len(rows),
                currency=context.currency,
                total=f"{total:.2f}",
            ),
            document=buf.getvalue(),
            document_name=fname,
        )

    async def _export_tasks(self, context: SessionContext) -> SkillResult:
        from sqlalchemy import text

        from src.core.db import async_session

        async with async_session() as session:
            result = await session.execute(
                text("""
                    SELECT title, status, deadline, created_at, description
                    FROM tasks
                    WHERE family_id = :fid
                    ORDER BY created_at DESC
                """),
                {"fid": context.family_id},
            )
            rows = result.all()

        if not rows:
            return SkillResult(
                response_text=t_cached(
                    _STRINGS,
                    "no_tasks",
                    context.language or "en",
                    namespace="export_excel",
                )
            )

        wb = Workbook()
        ws = wb.active
        ws.title = t_cached(
            _STRINGS,
            "tasks_sheet",
            context.language or "en",
            namespace="export_excel",
        )
        ws.append(
            [
                t_cached(_STRINGS, "title", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "status", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "due_date", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "created", context.language or "en", namespace="export_excel"),
                t_cached(
                    _STRINGS, "description", context.language or "en", namespace="export_excel"
                ),
            ]
        )
        _style_headers(ws)

        for r in rows:
            ws.append(
                [
                    r.title or "",
                    r.status or "",
                    str(r.deadline) if r.deadline else "",
                    str(r.created_at) if r.created_at else "",
                    r.description or "",
                ]
            )
        _auto_fit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)

        return SkillResult(
            response_text=t_cached(
                _STRINGS,
                "exported_tasks",
                context.language or "en",
                namespace="export_excel",
                count=len(rows),
            ),
            document=buf.getvalue(),
            document_name="tasks.xlsx",
        )

    async def _export_contacts(self, context: SessionContext) -> SkillResult:
        from sqlalchemy import text

        from src.core.db import async_session

        async with async_session() as session:
            result = await session.execute(
                text("""
                    SELECT name, phone, email, role, notes
                    FROM contacts
                    WHERE family_id = :fid
                    ORDER BY name
                """),
                {"fid": context.family_id},
            )
            rows = result.all()

        if not rows:
            return SkillResult(
                response_text=t_cached(
                    _STRINGS,
                    "no_contacts",
                    context.language or "en",
                    namespace="export_excel",
                )
            )

        wb = Workbook()
        ws = wb.active
        ws.title = t_cached(
            _STRINGS,
            "contacts_sheet",
            context.language or "en",
            namespace="export_excel",
        )
        ws.append(
            [
                t_cached(_STRINGS, "name", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "phone", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "email", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "role", context.language or "en", namespace="export_excel"),
                t_cached(_STRINGS, "notes", context.language or "en", namespace="export_excel"),
            ]
        )
        _style_headers(ws)

        for r in rows:
            ws.append(
                [
                    r.name or "",
                    r.phone or "",
                    r.email or "",
                    r.role or "",
                    r.notes or "",
                ]
            )
        _auto_fit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)

        return SkillResult(
            response_text=t_cached(
                _STRINGS,
                "exported_contacts",
                context.language or "en",
                namespace="export_excel",
                count=len(rows),
            ),
            document=buf.getvalue(),
            document_name="contacts.xlsx",
        )

    def get_system_prompt(self, context: SessionContext) -> str:
        prompts = load_prompt(Path(__file__).parent)
        template = prompts.get("system_prompt", _DEFAULT_SYSTEM_PROMPT)
        return template.format(language=context.language)


skill = ExportExcelSkill()
