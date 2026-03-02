"""Fill DOCX/XLSX templates with user data."""

import asyncio
import io
import logging
import re
import uuid as uuid_mod
from typing import Any

from sqlalchemy import select

from src.core.context import SessionContext
from src.core.db import async_session
from src.core.models.document import Document
from src.core.models.enums import DocumentType
from src.core.observability import observe
from src.gateway.types import IncomingMessage
from src.skills._i18n import register_strings, t_cached
from src.skills.base import SkillResult
from src.tools.storage import upload_document

logger = logging.getLogger(__name__)

SYSTEM_PROMPT = """\
You fill document templates (DOCX/XLSX) with user data.
Upload a template with placeholders like {{name}}, {{date}}, {{amount}}.
Be concise. Use HTML tags for Telegram.
Respond in: {language}."""

# Regex to find Jinja2-style placeholders: {{ variable_name }}
PLACEHOLDER_RE = re.compile(r"\{\{\s*(\w+)\s*\}\}")

_STRINGS = {
    "en": {
        "no_file": (
            "Please upload a <b>DOCX</b> or <b>XLSX</b> template file.\n"
            "Use placeholders like <code>{{name}}</code>, <code>{{date}}</code>, "
            "<code>{{amount}}</code> in your template."
        ),
        "unsupported_format": (
            "Unsupported format. Please upload a <b>.docx</b> or <b>.xlsx</b> template."
        ),
        "fill_failed": "Failed to process the template. Make sure it's a valid file.",
        "filled_ok": "<b>Template filled</b> ({format})",
        "filled_fields": "Filled: {fields}",
        "unfilled_fields": "Not filled (no data): {fields}",
        "no_placeholders": "No placeholders found in the template.",
        "no_templates": (
            "No saved templates. Upload a DOCX/XLSX and say <b>save as template</b>."
        ),
        "templates_header": "<b>Your templates:</b>",
        "template_saved": (
            "Template <b>{name}</b> saved. Use <b>list templates</b> to see all."
        ),
        "template_not_found": "Template <b>{name}</b> not found.",
        "template_deleted": "Template <b>{name}</b> deleted.",
        "ask_template_name": "Which template? Say <b>delete template {name}</b>.",
    },
    "ru": {
        "no_file": (
            "Загрузите файл-шаблон <b>DOCX</b> или <b>XLSX</b>.\n"
            "Используйте заполнители вроде <code>{{name}}</code>, <code>{{date}}</code>, "
            "<code>{{amount}}</code> в шаблоне."
        ),
        "unsupported_format": (
            "Формат не поддерживается. Загрузите шаблон <b>.docx</b> или <b>.xlsx</b>."
        ),
        "fill_failed": "Не удалось обработать шаблон. Убедитесь, что файл корректен.",
        "filled_ok": "<b>Шаблон заполнен</b> ({format})",
        "filled_fields": "Заполнено: {fields}",
        "unfilled_fields": "Не заполнено (нет данных): {fields}",
        "no_placeholders": "В шаблоне не найдено заполнителей.",
        "no_templates": (
            "Нет сохранённых шаблонов. Загрузите DOCX/XLSX и скажите <b>сохрани как шаблон</b>."
        ),
        "templates_header": "<b>Ваши шаблоны:</b>",
        "template_saved": (
            "Шаблон <b>{name}</b> сохранён. Скажите <b>мои шаблоны</b> для просмотра."
        ),
        "template_not_found": "Шаблон <b>{name}</b> не найден.",
        "template_deleted": "Шаблон <b>{name}</b> удалён.",
        "ask_template_name": "Какой шаблон? Скажите <b>удали шаблон {name}</b>.",
    },
    "es": {
        "no_file": (
            "Suba un archivo de plantilla <b>DOCX</b> o <b>XLSX</b>.\n"
            "Use marcadores como <code>{{name}}</code>, <code>{{date}}</code>, "
            "<code>{{amount}}</code> en su plantilla."
        ),
        "unsupported_format": (
            "Formato no compatible. Suba una plantilla <b>.docx</b> o <b>.xlsx</b>."
        ),
        "fill_failed": "No se pudo procesar la plantilla. Verifique que el archivo sea valido.",
        "filled_ok": "<b>Plantilla completada</b> ({format})",
        "filled_fields": "Completados: {fields}",
        "unfilled_fields": "Sin datos: {fields}",
        "no_placeholders": "No se encontraron marcadores en la plantilla.",
        "no_templates": (
            "No hay plantillas guardadas. Suba un DOCX/XLSX y diga <b>guardar como plantilla</b>."
        ),
        "templates_header": "<b>Sus plantillas:</b>",
        "template_saved": (
            "Plantilla <b>{name}</b> guardada. Use <b>mis plantillas</b> para ver todas."
        ),
        "template_not_found": "Plantilla <b>{name}</b> no encontrada.",
        "template_deleted": "Plantilla <b>{name}</b> eliminada.",
        "ask_template_name": "Cual plantilla? Diga <b>eliminar plantilla {name}</b>.",
    },
}
register_strings("fill_template", _STRINGS)


def _extract_template_name(text: str, action: str) -> str:
    """Extract template name from phrases like 'save as template Invoice' or
    'delete template Invoice'."""
    patterns = [
        r"(?:save|сохрани)\s+(?:as\s+)?(?:template|шаблон)\s+(.+)",
        r"(?:delete|удали)\s+(?:template|шаблон)\s+(.+)",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return ""


def _build_context_data(context: SessionContext, intent_data: dict[str, Any]) -> dict[str, str]:
    """Build a dict of available data from user profile and context."""
    data: dict[str, str] = {}

    # Profile fields
    if context.user_profile:
        profile = context.user_profile
        for field in ("name", "first_name", "last_name", "email", "phone", "company", "address"):
            val = getattr(profile, field, None) or profile.__dict__.get(field)
            if val:
                data[field] = str(val)

    # Context fields
    if context.currency:
        data["currency"] = context.currency
    if context.language:
        data["language"] = context.language
    if context.timezone:
        data["timezone"] = context.timezone

    # Any extra values the user passed via intent_data
    extra = intent_data.get("template_values") or {}
    if isinstance(extra, dict):
        data.update({k: str(v) for k, v in extra.items() if v})

    return data


def _fill_docx(file_bytes: bytes, data: dict[str, str]) -> tuple[bytes, list[str], list[str]]:
    """Fill a DOCX template using docxtpl. Returns (output_bytes, filled, unfilled)."""
    from docxtpl import DocxTemplate

    doc = DocxTemplate(io.BytesIO(file_bytes))

    # Get all undeclared variables (placeholders) from the template
    placeholders = doc.get_undeclared_template_variables()

    filled = []
    unfilled = []
    fill_context: dict[str, str] = {}

    for ph in sorted(placeholders):
        key = ph.lower()
        if key in data:
            fill_context[ph] = data[key]
            filled.append(ph)
        else:
            fill_context[ph] = f"[{ph}]"
            unfilled.append(ph)

    doc.render(fill_context)

    output = io.BytesIO()
    doc.save(output)
    return output.getvalue(), filled, unfilled


def _fill_xlsx(file_bytes: bytes, data: dict[str, str]) -> tuple[bytes, list[str], list[str]]:
    """Fill an XLSX template by replacing {{placeholder}} patterns in cells."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes))

    filled = []
    unfilled = []
    seen: set[str] = set()

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    matches = PLACEHOLDER_RE.findall(cell.value)
                    for match in matches:
                        key = match.lower()
                        if key not in seen:
                            seen.add(key)
                            if key in data:
                                filled.append(match)
                            else:
                                unfilled.append(match)
                        if key in data:
                            cell.value = cell.value.replace("{{" + match + "}}", data[key])
                            # Also handle spaced variant {{ match }}
                            cell.value = cell.value.replace("{{ " + match + " }}", data[key])

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue(), filled, unfilled


class FillTemplateSkill:
    name = "fill_template"
    intents = ["fill_template"]
    model = "claude-sonnet-4-6"

    @observe(name="fill_template")
    async def execute(
        self,
        message: IncomingMessage,
        context: SessionContext,
        intent_data: dict[str, Any],
    ) -> SkillResult:
        lang = context.language or "en"
        file_bytes = message.document_bytes
        filename = message.document_file_name or ""
        text = (message.text or "").strip().lower()

        # Derive extension early so template-save commands can use it
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        # --- Template library commands ---

        if "list template" in text or "мои шаблон" in text or "my template" in text:
            return await self._list_templates(context, lang)

        if "delete template" in text or "удали шаблон" in text:
            template_name = intent_data.get("template_name") or _extract_template_name(
                text, "delete"
            )
            return await self._delete_template(context, template_name, lang)

        # Save template: file attached + user asks to save as template
        if (
            file_bytes
            and ("save" in text or "сохрани" in text)
            and ("template" in text or "шаблон" in text)
        ):
            template_name = intent_data.get("template_name") or _extract_template_name(text, "save")
            return await self._save_template(
                context, file_bytes, filename, ext, template_name, lang
            )

        # --- Existing fill-template logic ---

        if not file_bytes:
            return SkillResult(response_text=t_cached(_STRINGS, "no_file", lang, "fill_template"))

        if ext not in ("docx", "xlsx"):
            return SkillResult(
                response_text=t_cached(
                    _STRINGS, "unsupported_format", lang, "fill_template"
                )
            )

        data = _build_context_data(context, intent_data)

        try:
            if ext == "docx":
                output_bytes, filled, unfilled = await asyncio.to_thread(
                    _fill_docx, file_bytes, data
                )
            else:
                output_bytes, filled, unfilled = await asyncio.to_thread(
                    _fill_xlsx, file_bytes, data
                )
        except Exception as e:
            logger.exception("Template fill failed: %s", e)
            return SkillResult(
                response_text=t_cached(_STRINGS, "fill_failed", lang, "fill_template")
            )

        # Build response text
        fmt = ext.upper()
        parts = [
            t_cached(_STRINGS, "filled_ok", lang, "fill_template").format(format=fmt)
        ]
        if filled:
            fields_str = ", ".join(f"<code>{f}</code>" for f in filled)
            parts.append(
                t_cached(_STRINGS, "filled_fields", lang, "fill_template").format(
                    fields=fields_str
                )
            )
        if unfilled:
            fields_str = ", ".join(f"<code>{u}</code>" for u in unfilled)
            parts.append(
                t_cached(_STRINGS, "unfilled_fields", lang, "fill_template").format(
                    fields=fields_str
                )
            )
        if not filled and not unfilled:
            parts.append(
                t_cached(_STRINGS, "no_placeholders", lang, "fill_template")
            )

        output_name = filename.replace(f".{ext}", f"_filled.{ext}")
        return SkillResult(
            response_text="\n".join(parts),
            document=output_bytes,
            document_name=output_name,
        )

    async def _list_templates(self, context: SessionContext, lang: str) -> SkillResult:
        """List user's saved templates."""
        async with async_session() as session:
            result = await session.execute(
                select(Document)
                .where(Document.family_id == uuid_mod.UUID(context.family_id))
                .where(Document.type == DocumentType.template)
                .order_by(Document.created_at.desc())
                .limit(20)
            )
            templates = result.scalars().all()

        if not templates:
            return SkillResult(
                response_text=t_cached(_STRINGS, "no_templates", lang, "fill_template")
            )

        lines = [t_cached(_STRINGS, "templates_header", lang, "fill_template")]
        for t in templates:
            name = (t.metadata_extra or {}).get("template_name", t.file_name or "Untitled")
            lines.append(f"  • {name} ({t.file_name})")
        return SkillResult(response_text="\n".join(lines))

    async def _save_template(
        self,
        context: SessionContext,
        file_bytes: bytes,
        filename: str,
        ext: str,
        template_name: str,
        lang: str,
    ) -> SkillResult:
        """Save a file as a reusable template."""
        if not template_name:
            template_name = filename.rsplit(".", 1)[0] if filename else "Untitled"

        if ext == "docx":
            mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        else:
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # Upload file bytes to Supabase Storage
        storage_path = await upload_document(
            file_bytes, context.family_id, filename, mime_type=mime, bucket="documents"
        )

        doc = Document(
            family_id=uuid_mod.UUID(context.family_id),
            user_id=uuid_mod.UUID(context.user_id),
            type=DocumentType.template,
            storage_path=storage_path,
            file_name=filename,
            title=template_name,
            mime_type=mime,
            file_size_bytes=len(file_bytes),
            metadata_extra={"template_name": template_name, "format": ext},
        )

        async with async_session() as session:
            session.add(doc)
            await session.commit()

        return SkillResult(
            response_text=t_cached(
                _STRINGS, "template_saved", lang, "fill_template"
            ).format(name=template_name)
        )

    async def _delete_template(
        self, context: SessionContext, template_name: str, lang: str
    ) -> SkillResult:
        """Delete a saved template by name."""
        if not template_name:
            return SkillResult(
                response_text=t_cached(
                    _STRINGS, "ask_template_name", lang, "fill_template"
                )
            )

        async with async_session() as session:
            result = await session.execute(
                select(Document)
                .where(Document.family_id == uuid_mod.UUID(context.family_id))
                .where(Document.type == DocumentType.template)
                .where(Document.title.ilike(f"%{template_name}%"))
                .limit(1)
            )
            doc = result.scalar_one_or_none()
            if not doc:
                return SkillResult(
                    response_text=t_cached(
                        _STRINGS, "template_not_found", lang, "fill_template"
                    ).format(name=template_name)
                )
            await session.delete(doc)
            await session.commit()

        return SkillResult(
            response_text=t_cached(
                _STRINGS, "template_deleted", lang, "fill_template"
            ).format(name=template_name)
        )

    def get_system_prompt(self, context: SessionContext) -> str:
        lang = context.language or "en"
        return SYSTEM_PROMPT.format(language=lang)


skill = FillTemplateSkill()
